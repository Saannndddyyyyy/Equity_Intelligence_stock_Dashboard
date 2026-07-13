import io
import logging
from datetime import datetime, date
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db, engine, Base
from app.models.watchlist import Watchlist
from app.models.alert import Alert
from app.models.recommendation import AIRecommendation
from app.services.portfolio import PortfolioService
from app.services.market_data import MarketDataService
from app.scheduler.tasks import start_scheduler, scheduler

# Configure Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Equity Stock Monitoring System",
    description="Investment Intelligence Dashboard for Indian Equities",
    version="1.0.0",
)

# Enable CORS for frontend flexibility
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Startup Hook to initialize scheduler
@app.on_event("startup")
def startup_event():
    logger.info("Application booting up. Initializing background tasks...")
    # Initialize DB tables if they don't exist
    Base.metadata.create_all(bind=engine)
    start_scheduler()


@app.on_event("shutdown")
def shutdown_event():
    logger.info("Application shutting down. Stopping scheduler...")
    if scheduler.running:
        scheduler.shutdown()


@app.get("/health", status_code=200)
def health_check():
    """
    Standard health check endpoint.
    """
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "scheduler_running": scheduler.running,
    }


# ----------------- WATCHLIST CRUD ENDPOINTS -----------------


@app.get("/api/watchlist")
def get_watchlist(db: Session = Depends(get_db)):
    return db.query(Watchlist).order_by(Watchlist.symbol.asc()).all()


@app.post("/api/watchlist")
def add_watchlist_item(item: dict, db: Session = Depends(get_db)):
    symbol = item.get("symbol", "").strip().upper()
    if not symbol:
        raise HTTPException(status_code=400, detail="Symbol field is required.")

    # Check if duplicate
    existing = db.query(Watchlist).filter(Watchlist.symbol == symbol).first()
    if existing:
        raise HTTPException(
            status_code=400, detail=f"Stock {symbol} is already in the watchlist."
        )

    # Auto-resolve name using MarketDataService
    market_service = MarketDataService()
    try:
        info = market_service.get_stock_info(symbol)
        company_name = info.get("company_name", symbol)
    except Exception:
        company_name = item.get("company_name") or symbol

    # Date parser
    p_date = None
    if item.get("purchase_date"):
        try:
            p_date = datetime.strptime(item["purchase_date"], "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Purchase date must be in YYYY-MM-DD format."
            )

    db_item = Watchlist(
        symbol=symbol,
        exchange=item.get("exchange", "NSE").strip().upper(),
        company_name=company_name,
        purchase_price=(
            float(item["purchase_price"])
            if item.get("purchase_price") is not None
            else None
        ),
        quantity=int(item["quantity"]) if item.get("quantity") is not None else 0,
        average_cost=(
            float(item["average_cost"])
            if item.get("average_cost") is not None
            else None
        ),
        target_price=(
            float(item["target_price"])
            if item.get("target_price") is not None
            else None
        ),
        stop_loss=(
            float(item["stop_loss"]) if item.get("stop_loss") is not None else None
        ),
        purchase_date=p_date,
        notes=item.get("notes", ""),
    )

    db.add(db_item)
    try:
        db.commit()
        db.refresh(db_item)
        return db_item
    except Exception as e:
        db.rollback()
        if "UNIQUE constraint" in str(e) or "duplicate key" in str(e):
            raise HTTPException(status_code=400, detail=f"The stock symbol '{symbol}' is already present in your watchlist.")
        raise HTTPException(status_code=500, detail="A database error occurred while saving the stock. Please try again.")


@app.put("/api/watchlist/{item_id}")
def update_watchlist_item(item_id: int, item: dict, db: Session = Depends(get_db)):
    db_item = db.query(Watchlist).filter(Watchlist.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Watchlist item not found.")

    if item.get("purchase_date"):
        try:
            db_item.purchase_date = datetime.strptime(
                item["purchase_date"], "%Y-%m-%d"
            ).date()
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Purchase date must be in YYYY-MM-DD format."
            )
    else:
        db_item.purchase_date = None

    db_item.exchange = item.get("exchange", db_item.exchange).strip().upper()
    db_item.purchase_price = (
        float(item["purchase_price"])
        if item.get("purchase_price") is not None
        else None
    )
    db_item.quantity = int(item["quantity"]) if item.get("quantity") is not None else 0
    db_item.average_cost = (
        float(item["average_cost"]) if item.get("average_cost") is not None else None
    )
    db_item.target_price = (
        float(item["target_price"]) if item.get("target_price") is not None else None
    )
    db_item.stop_loss = (
        float(item["stop_loss"]) if item.get("stop_loss") is not None else None
    )
    db_item.notes = item.get("notes", db_item.notes)

    try:
        db.commit()
        db.refresh(db_item)
        return db_item
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="A database error occurred while updating the watchlist details. Please try again.")


@app.delete("/api/watchlist/{item_id}")
def delete_watchlist_item(item_id: int, db: Session = Depends(get_db)):
    db_item = db.query(Watchlist).filter(Watchlist.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Watchlist item not found.")

    db.delete(db_item)
    db.commit()
    return {"message": f"Successfully removed {db_item.symbol} from watchlist."}


# ----------------- PORTFOLIO & AUDIT ENDPOINTS -----------------


@app.get("/api/portfolio/summary")
def get_portfolio_summary_api(db: Session = Depends(get_db)):
    p_service = PortfolioService(db)
    return p_service.get_portfolio_summary()


@app.get("/api/alerts")
def get_alerts(db: Session = Depends(get_db)):
    return db.query(Alert).order_by(Alert.sent_at.desc()).limit(50).all()


@app.get("/api/recommendations")
def get_recommendations(db: Session = Depends(get_db)):
    return (
        db.query(AIRecommendation)
        .order_by(AIRecommendation.created_at.desc())
        .limit(50)
        .all()
    )


# ----------------- EXCEL TEMPLATE IMPORT / EXPORT -----------------

TEMPLATE_HEADERS = [
    "Symbol",
    "Exchange",
    "Company Name",
    "Purchase Price",
    "Quantity",
    "Average Cost",
    "Target Price",
    "Stop Loss",
    "Purchase Date (YYYY-MM-DD)",
    "Notes",
]


@app.get("/api/watchlist/download-template")
def download_template():
    """
    Generates a blank Excel template file with headers and an example row.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Watchlist Template"

    # Style header row
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws.append(TEMPLATE_HEADERS)
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Add one example row
    example_row = [
        "RELIANCE",
        "NSE",
        "Reliance Industries Limited",
        1250.00,
        10,
        1250.00,
        1500.00,
        1150.00,
        "2026-07-13",
        "Growth catalyst: oil to retail expansion",
    ]
    ws.append(example_row)

    # Adjust columns width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

    # Save to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {"Content-Disposition": 'attachment; filename="watchlist_template.xlsx"'}
    return StreamingResponse(
        output,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/watchlist/upload-template")
def upload_template(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Reads a filled watchlist Excel template and validates cell entries before importing.
    Identifies specific cell and row indices for errors.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload a standard Excel file.",
        )

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")

    # Validate header column structure
    row_generator = ws.iter_rows(values_only=True)
    headers = next(row_generator, None)
    if not headers:
        raise HTTPException(status_code=400, detail="Excel file is empty.")

    # Match header columns
    for idx, expected in enumerate(TEMPLATE_HEADERS):
        if (
            idx >= len(headers)
            or str(headers[idx]).strip().lower()
            != expected.split(" (")[0].strip().lower()
        ):
            raise HTTPException(
                status_code=400,
                detail=f"Template header structure mismatch. Expected column #{idx+1} to be '{expected}', found '{headers[idx] if idx < len(headers) else 'None'}'.",
            )

    imported_items = []
    errors = []

    # Row 1 is header, data starts at Row 2
    for row_num, row_data in enumerate(row_generator, start=2):
        # Skip fully empty rows
        if all(val is None or str(val).strip() == "" for val in row_data):
            continue

        symbol = str(row_data[0] or "").strip().upper()
        exchange = str(row_data[1] or "NSE").strip().upper()
        company_name = str(row_data[2] or "").strip()
        purchase_price = row_data[3]
        quantity = row_data[4]
        average_cost = row_data[5]
        target_price = row_data[6]
        stop_loss = row_data[7]
        purchase_date_val = row_data[8]
        notes = str(row_data[9] or "").strip()

        # Validation checks
        if not symbol:
            errors.append(f"Row {row_num}: Column 'Symbol' is blank.")
            continue

        if purchase_price is not None:
            try:
                purchase_price = float(purchase_price)
            except ValueError:
                errors.append(
                    f"Row {row_num}: Purchase Price value '{purchase_price}' is not a valid decimal."
                )
                continue

        if quantity is not None:
            try:
                quantity = int(quantity)
            except ValueError:
                errors.append(
                    f"Row {row_num}: Quantity value '{quantity}' is not a valid integer."
                )
                continue

        if average_cost is not None:
            try:
                average_cost = float(average_cost)
            except ValueError:
                errors.append(
                    f"Row {row_num}: Average Cost value '{average_cost}' is not a valid decimal."
                )
                continue

        if target_price is not None:
            try:
                target_price = float(target_price)
            except ValueError:
                errors.append(
                    f"Row {row_num}: Target Price value '{target_price}' is not a valid decimal."
                )
                continue

        if stop_loss is not None:
            try:
                stop_loss = float(stop_loss)
            except ValueError:
                errors.append(
                    f"Row {row_num}: Stop Loss value '{stop_loss}' is not a valid decimal."
                )
                continue

        parsed_date = None
        if purchase_date_val is not None:
            # openpyxl might parse dates directly as datetime objects or string formats
            if isinstance(purchase_date_val, datetime):
                parsed_date = purchase_date_val.date()
            elif isinstance(purchase_date_val, date):
                parsed_date = purchase_date_val
            else:
                try:
                    parsed_date = datetime.strptime(
                        str(purchase_date_val).strip(), "%Y-%m-%d"
                    ).date()
                except ValueError:
                    errors.append(
                        f"Row {row_num}: Purchase Date value '{purchase_date_val}' is not in YYYY-MM-DD format."
                    )
                    continue

        # Check for symbol conflicts in DB
        existing = db.query(Watchlist).filter(Watchlist.symbol == symbol).first()
        if existing:
            errors.append(
                f"Row {row_num}: Symbol '{symbol}' is already present in your watchlist database."
            )
            continue

        imported_items.append(
            {
                "symbol": symbol,
                "exchange": exchange,
                "company_name": company_name or symbol,
                "purchase_price": purchase_price,
                "quantity": quantity or 0,
                "average_cost": average_cost,
                "target_price": target_price,
                "stop_loss": stop_loss,
                "purchase_date": parsed_date,
                "notes": notes,
            }
        )

    # If any error, reject the whole upload to ensure clean data loading
    if errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Upload rejected due to data validation errors.",
                "errors": errors,
            },
        )

    # Save to Database
    inserted = []
    for item in imported_items:
        db_item = Watchlist(**item)
        db.add(db_item)
        inserted.append(item["symbol"])

    try:
        db.commit()
        return {
            "message": f"Successfully imported {len(inserted)} items from template.",
            "symbols": inserted,
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail="A database transaction error occurred during template import. Please try again.",
        )


# ----------------- WEB DASHBOARD RENDERER -----------------


@app.get("/", response_class=HTMLResponse)
def index_page():
    return """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Investment Intelligence Stock Monitoring System</title>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
        <style>
            :root {
                --primary: #2563eb;
                --primary-hover: #1d4ed8;
                --background: #0f172a;
                --card-bg: #1e293b;
                --border: #334155;
                --text: #f8fafc;
                --text-muted: #94a3b8;
                --success: #10b981;
                --danger: #ef4444;
                --warning: #f59e0b;
                --neutral: #64748b;
            }

            * {
                box-sizing: border-box;
                margin: 0;
                padding: 0;
            }

            body {
                font-family: 'Inter', sans-serif;
                background-color: var(--background);
                color: var(--text);
                line-height: 1.5;
                padding-bottom: 50px;
            }

            header {
                background-color: var(--card-bg);
                border-bottom: 1px solid var(--border);
                padding: 20px 40px;
                display: flex;
                justify-content: space-between;
                align-items: center;
                position: sticky;
                top: 0;
                z-index: 10;
            }

            header h1 {
                font-size: 1.5rem;
                font-weight: 700;
                background: linear-gradient(to right, #60a5fa, #2563eb);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }

            .btn {
                background-color: var(--primary);
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                cursor: pointer;
                font-weight: 500;
                font-size: 0.875rem;
                transition: background-color 0.2s, transform 0.1s;
                display: inline-flex;
                align-items: center;
                gap: 8px;
            }

            .btn:hover {
                background-color: var(--primary-hover);
            }

            .btn-secondary {
                background-color: transparent;
                border: 1px solid var(--border);
                color: var(--text);
            }

            .btn-secondary:hover {
                background-color: #334155;
            }

            .btn-danger {
                background-color: var(--danger);
            }
            
            .btn-danger:hover {
                background-color: #b91c1c;
            }

            .container {
                max-width: 1400px;
                margin: 40px auto 0 auto;
                padding: 0 20px;
                display: grid;
                grid-template-columns: 3fr 1fr;
                gap: 30px;
            }

            @media (max-width: 1024px) {
                .container {
                    grid-template-columns: 1fr;
                }
            }

            .card {
                background-color: var(--card-bg);
                border: 1px solid var(--border);
                border-radius: 12px;
                padding: 24px;
                margin-bottom: 30px;
                box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.3);
            }

            .card-title {
                font-size: 1.125rem;
                font-weight: 600;
                margin-bottom: 20px;
                display: flex;
                justify-content: space-between;
                align-items: center;
                border-bottom: 1px solid var(--border);
                padding-bottom: 10px;
            }

            .portfolio-grid {
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 20px;
                margin-bottom: 30px;
            }

            @media (max-width: 768px) {
                .portfolio-grid {
                    grid-template-columns: repeat(2, 1fr);
                }
            }

            .stat-box {
                background-color: #0f172a;
                border: 1px solid var(--border);
                border-radius: 8px;
                padding: 16px;
                text-align: center;
            }

            .stat-val {
                font-size: 1.5rem;
                font-weight: 700;
                margin-top: 4px;
            }

            /* Tables */
            table {
                width: 100%;
                border-collapse: collapse;
                text-align: left;
                font-size: 0.875rem;
            }

            th, td {
                padding: 12px 16px;
                border-bottom: 1px solid var(--border);
            }

            th {
                color: var(--text-muted);
                font-weight: 500;
                text-transform: uppercase;
                font-size: 0.75rem;
            }

            tr:hover {
                background-color: #334155;
            }

            .badge {
                padding: 2px 8px;
                border-radius: 4px;
                font-size: 0.75rem;
                font-weight: 600;
            }

            .badge-success { background-color: rgba(16, 185, 129, 0.2); color: var(--success); }
            .badge-danger { background-color: rgba(239, 68, 68, 0.2); color: var(--danger); }
            .badge-warning { background-color: rgba(245, 158, 11, 0.2); color: var(--warning); }
            .badge-neutral { background-color: rgba(100, 116, 139, 0.2); color: var(--neutral); }

            /* Forms */
            .form-grid {
                display: grid;
                grid-template-columns: repeat(3, 1fr);
                gap: 15px;
                margin-bottom: 15px;
            }

            @media (max-width: 768px) {
                .form-grid {
                    grid-template-columns: 1fr;
                }
            }

            .form-group {
                display: flex;
                flex-direction: column;
                gap: 6px;
                position: relative;
            }

            .form-group label {
                font-size: 0.75rem;
                font-weight: 500;
                color: var(--text-muted);
                display: flex;
                align-items: center;
                gap: 5px;
            }

            .form-group input, .form-group select, .form-group textarea {
                background-color: #0f172a;
                border: 1px solid var(--border);
                border-radius: 6px;
                padding: 8px 12px;
                color: white;
                font-size: 0.875rem;
            }

            .form-group input:focus, .form-group select:focus {
                outline: none;
                border-color: var(--primary);
            }

            /* Popovers */
            .info-icon {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 14px;
                height: 14px;
                background-color: var(--border);
                color: var(--text-muted);
                border-radius: 50%;
                font-size: 10px;
                cursor: pointer;
                font-style: normal;
                font-weight: bold;
            }

            .info-icon:hover::after {
                content: attr(data-tooltip);
                position: absolute;
                background-color: #1e293b;
                border: 1px solid var(--border);
                color: white;
                padding: 8px 12px;
                border-radius: 4px;
                font-size: 11px;
                width: 220px;
                white-space: normal;
                z-index: 100;
                top: -60px;
                left: 0;
                box-shadow: 0 4px 6px -1px rgba(0,0,0,0.5);
            }

            /* Modals */
            .modal {
                display: none;
                position: fixed;
                top: 0;
                left: 0;
                width: 100%;
                height: 100%;
                background-color: rgba(0,0,0,0.6);
                align-items: center;
                justify-content: center;
                z-index: 100;
            }

            .modal-content {
                background-color: var(--card-bg);
                border: 1px solid var(--border);
                border-radius: 12px;
                padding: 30px;
                width: 100%;
                max-width: 600px;
                box-shadow: 0 20px 25px -5px rgba(0,0,0,0.5);
                position: relative;
            }

            .close-modal {
                position: absolute;
                top: 15px;
                right: 20px;
                font-size: 1.5rem;
                cursor: pointer;
                color: var(--text-muted);
            }

            /* Alerts panel */
            .alert-item {
                border-left: 4px solid var(--neutral);
                padding: 10px 15px;
                margin-bottom: 12px;
                background-color: #0f172a;
                border-radius: 0 6px 6px 0;
            }

            .alert-item.high { border-left-color: var(--danger); }
            .alert-item.medium { border-left-color: var(--warning); }
            .alert-item.low { border-left-color: var(--primary); }
        </style>
    </head>
    <body>
        <header>
            <h1>Equity Intelligence stock Dashboard</h1>
            <div style="display: flex; gap: 10px;">
                <button class="btn btn-secondary" onclick="openModal('howToUseModal')">How to Use</button>
                <button class="btn" onclick="downloadTemplate()">Download Excel Template</button>
            </div>
        </header>

        <div class="container">
            <!-- Left Side main operations -->
            <div>
                <!-- Portfolio Valuation summary panel -->
                <div class="card">
                    <div class="card-title">Portfolio Intelligence Stats</div>
                    <div class="portfolio-grid">
                        <div class="stat-box">
                            <div style="font-size: 0.75rem; color: var(--text-muted);">Total Cost Value</div>
                            <div class="stat-val" id="val-cost">INR 0.00</div>
                        </div>
                        <div class="stat-box">
                            <div style="font-size: 0.75rem; color: var(--text-muted);">Market Valuation</div>
                            <div class="stat-val" id="val-market">INR 0.00</div>
                        </div>
                        <div class="stat-box">
                            <div style="font-size: 0.75rem; color: var(--text-muted);">Total PnL Returns</div>
                            <div class="stat-val" id="val-pnl" style="color: var(--success);">INR 0.00 (0.00%)</div>
                        </div>
                        <div class="stat-box">
                            <div style="font-size: 0.75rem; color: var(--text-muted);">Current Horizon State</div>
                            <div class="stat-val" style="color: var(--warning);" id="val-rules-month">Month 1 (Growth)</div>
                        </div>
                    </div>
                </div>

                <!-- Watchlist Table Section -->
                <div class="card">
                    <div class="card-title">
                        <span>Active Equity Watchlist</span>
                        <div style="display: flex; gap: 10px;">
                            <input type="file" id="upload-file" style="display: none;" onchange="handleExcelUpload(this)">
                            <button class="btn btn-secondary" onclick="document.getElementById('upload-file').click()">Upload Excel File</button>
                            <button class="btn" onclick="openModal('addStockModal')">Add Stock Symbol</button>
                        </div>
                    </div>
                    <table id="watchlist-table">
                        <thead>
                            <tr>
                                <th>Symbol</th>
                                <th>Exchange</th>
                                <th>Qty</th>
                                <th>Avg Cost</th>
                                <th>Target Price</th>
                                <th>Stop Loss</th>
                                <th>Purchase Date</th>
                                <th>Notes</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td colspan="9" style="text-align: center; color: var(--text-muted);">No stock symbols watched yet. Add symbols to start monitoring.</td>
                            </tr>
                        </tbody>
                    </table>
                </div>

                <!-- AI Recommendations history list -->
                <div class="card">
                    <div class="card-title">Historical AI Recommendations</div>
                    <div id="recommendations-container" style="display: flex; flex-direction: column; gap: 15px;">
                        <p style="color: var(--text-muted); font-size: 0.875rem;">No recommendations generated yet. High impact events trigger AI generation cycle.</p>
                    </div>
                </div>
            </div>

            <!-- Right side logs bar -->
            <div>
                <div class="card">
                    <div class="card-title">Event Alerts Feed</div>
                    <div id="alerts-container" style="max-height: 500px; overflow-y: auto;">
                        <p style="color: var(--text-muted); font-size: 0.875rem;">No alerts recorded.</p>
                    </div>
                </div>
            </div>
        </div>

        <!-- Add Stock Modal -->
        <div class="modal" id="addStockModal">
            <div class="modal-content">
                <span class="close-modal" onclick="closeModal('addStockModal')">&times;</span>
                <h3 style="margin-bottom: 20px;">Add Stock to watchlist</h3>
                <form id="add-stock-form" onsubmit="handleAddStock(event)">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Symbol <i class="info-icon" data-tooltip="The ticker symbol of the equity stock. (e.g. RELIANCE)">i</i></label>
                            <input type="text" id="form-symbol" placeholder="e.g. RELIANCE" required>
                        </div>
                        <div class="form-group">
                            <label>Exchange <i class="info-icon" data-tooltip="Trading exchange where the share is listed. NSE or BSE">i</i></label>
                            <select id="form-exchange">
                                <option value="NSE">NSE</option>
                                <option value="BSE">BSE</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Purchase Price <i class="info-icon" data-tooltip="Price per share at time of buy. (e.g. 1250.00)">i</i></label>
                            <input type="number" step="0.01" id="form-purchase-price" placeholder="e.g. 1250.00">
                        </div>
                    </div>
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Quantity <i class="info-icon" data-tooltip="Number of shares bought. (e.g. 10)">i</i></label>
                            <input type="number" id="form-qty" placeholder="e.g. 10">
                        </div>
                        <div class="form-group">
                            <label>Average Cost <i class="info-icon" data-tooltip="Average cost including commission charges. (e.g. 1250.00)">i</i></label>
                            <input type="number" step="0.01" id="form-avg-cost" placeholder="e.g. 1250.00">
                        </div>
                        <div class="form-group">
                            <label>Target Price <i class="info-icon" data-tooltip="Your exit sell target price. (e.g. 1500.00)">i</i></label>
                            <input type="number" step="0.01" id="form-target" placeholder="e.g. 1500.00">
                        </div>
                    </div>
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Stop Loss <i class="info-icon" data-tooltip="Your downside exit threshold trigger. (e.g. 1150.00)">i</i></label>
                            <input type="number" step="0.01" id="form-stop" placeholder="e.g. 1150.00">
                        </div>
                        <div class="form-group">
                            <label>Purchase Date <i class="info-icon" data-tooltip="Format YYYY-MM-DD. (e.g. 2026-07-13)">i</i></label>
                            <input type="text" id="form-date" placeholder="YYYY-MM-DD">
                        </div>
                    </div>
                    <div class="form-group" style="margin-bottom: 20px;">
                        <label>Notes <i class="info-icon" data-tooltip="Any additional research notes or reasons for watch.">i</i></label>
                        <textarea id="form-notes" rows="3" placeholder="e.g. Earnings growth catalyst..."></textarea>
                    </div>
                    <button class="btn" style="width: 100%; justify-content: center;" type="submit">Submit Watchlist Stock</button>
                </form>
            </div>
        </div>

        <!-- How to Use Modal -->
        <div class="modal" id="howToUseModal">
            <div class="modal-content" style="max-width: 700px;">
                <span class="close-modal" onclick="closeModal('howToUseModal')">&times;</span>
                <h3 style="margin-bottom: 20px; border-bottom: 1px solid var(--border); padding-bottom: 10px;">How to Use the Stock Monitoring System</h3>
                <div style="font-size: 0.9rem; line-height: 1.6; max-height: 400px; overflow-y: auto; padding-right: 10px;">
                    <p style="margin-bottom: 10px; font-weight: bold;">Follow these simple steps to manage your portfolio:</p>
                    <ol style="margin-left: 20px; margin-bottom: 20px; display: flex; flex-direction: column; gap: 8px;">
                        <li><strong>Add Stocks</strong>: Click the "Add Stock Symbol" button and fill in details like ticker (e.g. RELIANCE, TCS), targets, and notes.</li>
                        <li><strong>Bulk Upload</strong>: Click "Download Excel Template" to fetch a pre-formatted excel file, write your holding records, and upload it via "Upload Excel File".</li>
                        <li><strong>Track Performance</strong>: The "Portfolio Intelligence Stats" panel dynamically fetches live stock prices and prints returns and allocations.</li>
                        <li><strong>Monitor Alerts</strong>: The Right-hand sidebar lists price movements, technical volume spikes, and news events.</li>
                        <li><strong>AI Advice</strong>: Important events trigger AI recommendations tailored specifically for your <strong>9-month investment timeline</strong>, helping you buy in Months 1-5 and liquidate safely by Month 9.</li>
                    </ol>
                    <p style="font-weight: 500; color: var(--warning);">Note: For Indian markets (NSE/BSE), ticker symbols are looked up automatically via yfinance (e.g. "RELIANCE" maps to RELIANCE.NS on the NSE).</p>
                </div>
            </div>
        </div>

        <script>
            // Modal state
            function openModal(id) { document.getElementById(id).style.display = 'flex'; }
            function closeModal(id) { document.getElementById(id).style.display = 'none'; }

            // Download template
            function downloadTemplate() {
                window.open("/api/watchlist/download-template", "_blank");
            }

            // Fetch watchlist and reload table
            async function loadWatchlist() {
                try {
                    const response = await fetch("/api/watchlist");
                    const data = await response.json();
                    const tbody = document.querySelector("#watchlist-table tbody");
                    
                    if (data.length === 0) {
                        tbody.innerHTML = `<tr><td colspan="9" style="text-align: center; color: var(--text-muted);">No stock symbols watched yet. Add symbols to start monitoring.</td></tr>`;
                        return;
                    }

                    tbody.innerHTML = "";
                    data.forEach(item => {
                        const tr = document.createElement("tr");
                        tr.innerHTML = `
                            <td style="font-weight: bold;">${item.symbol}</td>
                            <td>${item.exchange}</td>
                            <td>${item.quantity || 0}</td>
                            <td>INR ${item.average_cost ? item.average_cost.toFixed(2) : 'N/A'}</td>
                            <td>INR ${item.target_price ? item.target_price.toFixed(2) : 'N/A'}</td>
                            <td>INR ${item.stop_loss ? item.stop_loss.toFixed(2) : 'N/A'}</td>
                            <td>${item.purchase_date || 'N/A'}</td>
                            <td style="font-size: 0.75rem; max-width: 150px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">${item.notes || ''}</td>
                            <td>
                                <button class="btn btn-secondary btn-danger" style="padding: 4px 8px; font-size: 11px;" onclick="deleteStock(${item.id})">Remove</button>
                            </td>
                        `;
                        tbody.appendChild(tr);
                    });
                } catch (e) {
                    console.error("Error loading watchlist:", e);
                }
            }

            // Fetch portfolio stats
            async function loadPortfolio() {
                try {
                    const response = await fetch("/api/portfolio/summary");
                    const data = await response.json();
                    
                    document.getElementById("val-cost").innerText = `INR ${data.total_cost_value.toLocaleString(undefined, {minimumFractionDigits:2, maximumFractionDigits:2})}`;
                    document.getElementById("val-market").innerText = `INR ${data.total_market_value.toLocaleString(undefined, {minimumFractionDigits:2, maximumFractionDigits:2})}`;
                    
                    const pnlVal = document.getElementById("val-pnl");
                    pnlVal.innerText = `INR ${data.total_pnl.toLocaleString(undefined, {minimumFractionDigits:2, maximumFractionDigits:2})} (${data.total_pnl_percent.toFixed(2)}%)`;
                    
                    if (data.total_pnl >= 0) {
                        pnlVal.style.color = "var(--success)";
                    } else {
                        pnlVal.style.color = "var(--danger)";
                    }
                } catch (e) {
                    console.error("Error loading portfolio stats:", e);
                }
            }

            // Fetch alerts log
            async function loadAlerts() {
                try {
                    const response = await fetch("/api/alerts");
                    const data = await response.json();
                    const container = document.getElementById("alerts-container");
                    
                    if (data.length === 0) {
                        container.innerHTML = `<p style="color: var(--text-muted); font-size: 0.875rem;">No alerts recorded.</p>`;
                        return;
                    }

                    container.innerHTML = "";
                    data.forEach(item => {
                        const sevClass = item.severity.toLowerCase();
                        const time = new Date(item.sent_at).toLocaleTimeString();
                        const div = document.createElement("div");
                        div.className = `alert-item ${sevClass}`;
                        div.innerHTML = `
                            <div style="display: flex; justify-content: space-between; font-size: 0.75rem; color: var(--text-muted);">
                                <strong>${item.symbol} - ${item.event_type}</strong>
                                <span>${time}</span>
                            </div>
                            <p style="font-size: 0.85rem; margin-top: 4px;">${item.summary}</p>
                        `;
                        container.appendChild(div);
                    });
                } catch (e) {
                    console.error("Error loading alerts:", e);
                }
            }

            // Fetch recommendations
            async function loadRecommendations() {
                try {
                    const response = await fetch("/api/recommendations");
                    const data = await response.json();
                    const container = document.getElementById("recommendations-container");
                    
                    if (data.length === 0) {
                        container.innerHTML = `<p style="color: var(--text-muted); font-size: 0.875rem;">No recommendations generated yet. High impact events trigger AI generation cycle.</p>`;
                        return;
                    }

                    container.innerHTML = "";
                    data.forEach(item => {
                        const dateStr = new Date(item.created_at).toLocaleDateString();
                        const recClass = item.recommendation.toLowerCase().includes("buy") ? "badge-success" : item.recommendation.toLowerCase().includes("sell") ? "badge-danger" : "badge-warning";
                        const div = document.createElement("div");
                        div.style.border = "1px solid var(--border)";
                        div.style.padding = "15px";
                        div.style.borderRadius = "8px";
                        div.style.backgroundColor = "#0f172a";
                        div.innerHTML = `
                            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
                                <strong>${item.symbol}</strong>
                                <span class="badge ${recClass}">${item.recommendation}</span>
                            </div>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin-bottom: 8px;">${item.executive_summary}</p>
                            <div style="display: flex; gap: 20px; font-size: 0.75rem; color: var(--text-muted);">
                                <span>Action: <strong style="color: white;">${item.suggested_action}</strong></span>
                                <span>Horizon exit: <strong style="color: white;">${item.exit_strategy}</strong></span>
                                <span>Date: ${dateStr}</span>
                            </div>
                        `;
                        container.appendChild(div);
                    });
                } catch (e) {
                    console.error("Error loading recommendations:", e);
                }
            }

            // Delete stock
            async function deleteStock(id) {
                if (confirm("Are you sure you want to remove this stock?")) {
                    await fetch(`/api/watchlist/${id}`, { method: "DELETE" });
                    loadAll();
                }
            }

            // Submit stock form
            async function handleAddStock(event) {
                event.preventDefault();
                const submitBtn = event.target.querySelector("button[type='submit']");
                if (submitBtn) {
                    submitBtn.disabled = true;
                    submitBtn.innerText = "Submitting...";
                }
                const item = {
                    symbol: document.getElementById("form-symbol").value,
                    exchange: document.getElementById("form-exchange").value,
                    purchase_price: document.getElementById("form-purchase-price").value || null,
                    quantity: document.getElementById("form-qty").value || null,
                    average_cost: document.getElementById("form-avg-cost").value || null,
                    target_price: document.getElementById("form-target").value || null,
                    stop_loss: document.getElementById("form-stop").value || null,
                    purchase_date: document.getElementById("form-date").value || null,
                    notes: document.getElementById("form-notes").value
                };

                try {
                    const response = await fetch("/api/watchlist", {
                        method: "POST",
                        headers: { "Content-Type": "application/json" },
                        body: JSON.stringify(item)
                    });
                    
                    const resData = await response.json();
                    if (!response.ok) {
                        alert(resData.detail || "Failed to add stock.");
                    } else {
                        closeModal("addStockModal");
                        document.getElementById("add-stock-form").reset();
                        loadAll();
                    }
                } catch (e) {
                    alert("Network connection error adding stock.");
                } finally {
                    if (submitBtn) {
                        submitBtn.disabled = false;
                        submitBtn.innerText = "Submit Watchlist Stock";
                    }
                }
            }

            // Bulk Excel Upload
            async function handleExcelUpload(input) {
                if (!input.files || input.files.length === 0) return;
                const file = input.files[0];
                const formData = new FormData();
                formData.append("file", file);

                try {
                    const response = await fetch("/api/watchlist/upload-template", {
                        method: "POST",
                        body: formData
                    });
                    const data = await response.json();
                    if (!response.ok) {
                        // Standard error alert displaying individual rows and columns
                        if (data.detail && data.detail.errors) {
                            alert(data.detail.message + "\\n\\n" + data.detail.errors.join("\\n"));
                        } else {
                            alert(data.detail || "Failed to upload file.");
                        }
                    } else {
                        alert(data.message);
                        loadAll();
                    }
                } catch (e) {
                    alert("Network connection error uploading template.");
                }
                input.value = ""; // Reset input file select
            }

            // Aggregate loader
            function loadAll() {
                loadWatchlist();
                loadPortfolio();
                loadAlerts();
                loadRecommendations();
            }

            // Init call
            window.onload = function() {
                loadAll();
                // Refresh updates every 30 seconds
                setInterval(loadAll, 30000);
            };
        </script>
    </body>
    </html>
    """
